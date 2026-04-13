"""One-off script to write DEV comments + Status into the 10042026 tab
of 'AI PDF Tools Bug Document (3).xlsx' after all OP_001..OP_030 fixes shipped.
Safe to re-run; overwrites Status (H) and DEV comments (I) columns only.
"""
import openpyxl
from openpyxl.styles import PatternFill, Alignment

path = "manual-tests/AI PDF Tools Bug Document (3).xlsx"

fixes = {
    "OP_001": ("Fixed", "Batch C+E (commits ad398aa + 798d83f). Shared client-side validator lib/client-file-validator.ts wired into 6 AI interfaces + processing-interface + table-extraction + pdf-compare. File-type / empty-file / size-limit errors now surface on drop, not after spinner. Upload flow now uniform across tools."),
    "OP_002": ("Fixed", "Batch F (commit 6932a77). Integrated Google Safe Browsing v4 via lib/url-validation.js checkUrlSafety(). GOOGLE_SAFE_BROWSING_API_KEY added to Vercel and verified live with testsafebrowsing.appspot.com -- now returns 'This URL is unsafe' 400 error."),
    "OP_003": ("Fixed", "Batch E (commit 798d83f). components/pdf-compare-interface.tsx now catches pdfjs PasswordException and InvalidPDFException and surfaces 'This PDF is password-protected' / 'Not a valid PDF' instead of generic 'Failed to read the modified PDF'."),
    "OP_004": ("Fixed", "Batch H (commit b82fc94). app/api/create-portal-session/route.ts now returns tailored messages: 'No portal configuration' -> service unavailable + contact email; 'No such customer' -> re-link subscription; null stripe_customer_id -> contact paula to link (enterprise manual-upgrade case)."),
    "OP_005": ("Fixed", "Batch E (commit 798d83f). components/table-extraction-interface.tsx now uses acceptFile helper with validateClientFile + user-plan size limit. Shows inline error before API call. Size limit label added under dropzone."),
    "OP_006": ("Fixed", "Batch D (commit e2b05b5). app/api/generate-resume/route.ts prompt rewritten: 'Use ONLY the information provided. Do NOT invent employers, titles, dates, metrics, skills, or achievements.' Removed 'create realistic metrics' instruction. Applies to both ATS optimizer build flow and /create-resume."),
    "OP_007": ("Fixed", "Batch D (commit e2b05b5). Email now required server-side (400 error if missing) AND client-side in both resume-builder-interface.tsx and ats-optimizer-interface.tsx. Regex validation on both."),
    "OP_008": ("Fixed", "Batch D (commit e2b05b5). Phone now validated when provided: digits/spaces/()/+/- only, 7-20 digits total. Server-side 400 error + client-side inline error + HTML pattern attribute."),
    "OP_009": ("Fixed", "Batch C (commit ad398aa). Blank/empty file now rejected on drop or select, not after processing. ats-optimizer-interface.tsx uses acceptFile helper with validateClientFile. Backend isBlankPdf check already added in Batch B (commit d56e246)."),
    "OP_010": ("Fixed", "Batch D (commit e2b05b5). Education and Skills now required in both resume-builder-interface.tsx and ats-optimizer-interface.tsx Build Resume flow. Button disabled until populated; server returns 400 if missing."),
    "OP_011": ("Fixed", "Batch D (commit e2b05b5). Server detects 'no skills', 'none', 'n/a', '-' in skills/certifications/languages fields via isEmptyIntent() helper and instructs AI to OMIT that section entirely. User's literal input honored."),
    "OP_012": ("Fixed", "Batch A+B (commits e68eee2 + d56e246). Shared lib/pdf-content-guard.ts detects binary streams, metadata-only, and image-only content BEFORE sending to OpenAI. app/api/smart-extraction/route.ts now uses guardPdfContent() with specific user messages. Pre-upload client validation added in Batch C."),
    "OP_013": ("Fixed", "Batch A+B. Same guardPdfContent() wired into app/api/chat-with-pdf/route.ts. Valid PDFs no longer falsely rejected. Pre-upload client validation added in Batch C."),
    "OP_014": ("Fixed", "Batch F (commit 6932a77). All 3 signup pages (EN/ES/BR) now detect Supabase data.user.identities.length === 0 response and show 'An account with this email already exists. Please log in or reset your password.' instead of the misleading success screen."),
    "OP_015": ("Fixed", "Batch F (commit 6932a77). Strong password now enforced on signup (EN/ES/BR) AND reset-password: requires uppercase, lowercase, digit, and special character in addition to 8-char minimum. Hint text added under password field."),
    "OP_016": ("Fixed", "Batch H (commit b82fc94). app/contact/page.tsx now surfaces the actual API error in the inline banner instead of generic 'Failed to send'. State includes errorMessage; user sees rate-limit messages, validation errors, or service status directly."),
    "OP_017": ("Fixed", "Batch G (commit 38a0186). app/api/translate-pdf/route.ts now chunks input at 3000-char paragraph boundaries and translates chunks in parallel (max 4 concurrent). Raises input cap from 12k to 40k chars. No more silent truncation by max_tokens output cap."),
    "OP_018": ("Fixed", "Batch A/B/C. app/api/translate-pdf/route.ts rejects blank files BEFORE calling iLoveAPI via strict isBlankPdf() check; client-side rejection on upload via validateClientFile. Specific message: 'This file appears to be empty. Please upload a PDF with content.'"),
    "OP_019": ("Improved", "Batch G (commit 38a0186). Parallel chunk translation cuts total latency by up to 4x on long documents. Small documents unchanged. Frontend progress bar unchanged -- still shows 5% -> 40% -> 90% during AI call."),
    "OP_020": ("Fixed", "Batch A+B. guardPdfContent() detects image-only PDFs via word-density (<20 word tokens) and PDF stream-pattern checks. Returns 422 with: 'This PDF appears to contain only images and no readable text. Please run OCR first.'"),
    "OP_021": ("Fixed", "Batch A+B. Strict isBlankPdf() check in app/api/pdf-summarizer/route.ts rejects blank PDFs with 400 status before calling OpenAI. guardPdfContent() blocks metadata-only content as secondary defense."),
    "OP_022": ("Fixed", "Batch A+B+I. Image-only photo PDFs caught by guardPdfContent word-density check. Summarizer returns 422 with OCR guidance. Batch I added lower temperature + strict prompt in related routes."),
    "OP_023": ("Fixed", "Batch I (commit 97cae00). components/question-generator-interface.tsx adds Download button next to Copy All. Produces <original-pdf>-questions.txt with all questions, options, answers, explanations. Labels translated EN/ES/BR."),
    "OP_024": ("Fixed", "Batch I (commit 97cae00). app/api/smart-extraction/route.ts: response_format json_object enforces valid JSON. temperature dropped from 0.1 to 0. Prompt hardened: 'Use EXACT strings from document. Do not paraphrase. Leave ambiguous fields out.'"),
    "OP_025": ("Fixed", "Batch K (commit 4a50d27). All 3 pricing pages (EN/ES/BR): Free plan gets 'Free Forever' badge matching other cards pt-10 layout. Equal card heights via min-h-[20px] placeholder. Contrast bumped (slate-400 -> slate-300). CTAs unified to 'Start 30-Day Free Trial'."),
    "OP_026": ("Fixed", "Batch J (commit 8b05064). Privacy Policy and Terms & Conditions links in all 3 footers (EN/ES/BR) and cookie-consent banner now have target='_blank' rel='noopener noreferrer'. User no longer loses form state when clicking through."),
    "OP_027": ("Fixed", "Batch J (commit 8b05064). Lucide Twitter icon replaced with inline X logo SVG in all 3 footers (EN/ES/BR). aria-labels updated from 'Twitter' to 'X'."),
    "OP_028": ("Fixed", "Batch K (commit 4a50d27). Blog '{N} articles' pill was a static <div>. Now an <a href='#articles'> with hover state + ArrowRight icon that smooth-scrolls to articles grid. Grid section has id='articles' + scroll-mt-20. Applied to all 3 blog pages (EN/ES/BR)."),
    "OP_029": ("Fixed", "Batch J (commit 8b05064). HCaptcha widget wrapped in <div className='flex justify-center'> in all 3 signup pages (EN/ES/BR). Captcha now horizontally centered."),
    "OP_030": ("Fixed", "Batch C (commit ad398aa). components/processing/file-dropzone.tsx now shows 'Max file size: 25MB/200MB/1GB' label under the dropzone (plan-aware via getSizeLimitLabel(userPlan)). Same label added to 7 standalone AI interfaces."),
}

wb = openpyxl.load_workbook(path)
ws = wb["10042026"]

green = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

updated = 0
for row in range(2, ws.max_row + 1):
    op_id = ws.cell(row=row, column=1).value
    if op_id in fixes:
        status, comment = fixes[op_id]
        ws.cell(row=row, column=8, value=status)
        ws.cell(row=row, column=9, value=comment)
        ws.cell(row=row, column=8).fill = green if status == "Fixed" else yellow
        ws.cell(row=row, column=9).alignment = Alignment(wrap_text=True, vertical="top")
        updated += 1

wb.save(path)
print(f"Updated {updated} rows in 10042026 tab.")
